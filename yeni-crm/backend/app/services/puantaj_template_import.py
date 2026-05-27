"""Toplu puantaj Excel şablonu — geri yükleme (import).

Kullanıcı tarafından doldurulmuş şablonu okur ve daily_entries tablosuna
upsert eder. İki sheet'i de okur:
  - 'Puantaj' — kurye × gün matrisi (Saat / Paket / Durum 3 satır)
  - 'Destek' — destek vardiyaları (tarih, kurye, restoran, saat, paket)

Durum kodları → cell_type eşlemesi:
  G → gelmedi
  R → raporlu
  Z → izin
  X → ihbarsiz
  D → destek (saat/paket dolu, kendi restoranı dışında)
  boş + saat/paket dolu → normal
  boş + saat/paket boş → empty (kayıt yok)

Returns:
  {
    "puantaj": { "inserted": int, "updated": int, "skipped": int, "errors": [...] },
    "destek": { "inserted": int, "updated": int, "skipped": int, "errors": [...] }
  }
"""
from __future__ import annotations

import io
import logging
from calendar import monthrange
from typing import Any

from openpyxl import load_workbook
from psycopg.rows import dict_row

from app.core.database import get_connection
from app.services.puantaj import upsert_cell


log = logging.getLogger(__name__)


STATUS_CODE_MAP = {
    "G": "gelmedi",
    "R": "raporlu",
    "Z": "izin",
    "X": "ihbarsiz",
}


def _lookup_personnel(person_code: str) -> dict | None:
    if not person_code:
        return None
    code = person_code.strip().upper()
    with get_connection() as conn:
        with conn.cursor(row_factory=dict_row) as cur:
            cur.execute(
                """
                SELECT id, person_code, full_name, assigned_restaurant_id, role
                FROM personnel
                WHERE UPPER(person_code) = %s
                LIMIT 1
                """,
                (code,),
            )
            return cur.fetchone()


def _lookup_restaurant(label: str) -> int | None:
    """Restoran 'brand / branch' veya 'brand' formatı, veya id."""
    if not label:
        return None
    s = label.strip()
    if not s:
        return None
    # ID mi?
    try:
        return int(s)
    except ValueError:
        pass
    with get_connection() as conn:
        with conn.cursor(row_factory=dict_row) as cur:
            # Önce tam eşleme dene
            cur.execute(
                """
                SELECT id, brand, branch FROM restaurants
                WHERE LOWER(TRIM(brand || ' / ' || COALESCE(branch, ''))) = LOWER(%s)
                   OR LOWER(TRIM(brand)) = LOWER(%s)
                """,
                (s, s),
            )
            row = cur.fetchone()
            if row:
                return int(row["id"])
            # Substring fallback
            cur.execute(
                """
                SELECT id FROM restaurants
                WHERE LOWER(brand) LIKE %s OR LOWER(COALESCE(branch, '')) LIKE %s
                LIMIT 1
                """,
                (f"%{s.lower()}%", f"%{s.lower()}%"),
            )
            row = cur.fetchone()
            return int(row["id"]) if row else None


def import_puantaj_template(
    file_bytes: bytes, period: str,
) -> dict[str, Any]:
    """Excel dosyasını oku ve daily_entries'e işle.

    period: 'YYYY-MM' — header info için doğrulama amacı (gün şablonu eşleşmeli).
    file_bytes: yüklenen .xlsx bytes
    """
    yyyy_i, mm_i = [int(x) for x in period.split("-")]
    n_days = monthrange(yyyy_i, mm_i)[1]

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"Geçersiz Excel dosyası: {e}") from e

    result = {
        "puantaj": {"inserted": 0, "updated": 0, "skipped": 0, "errors": []},
        "destek": {"inserted": 0, "updated": 0, "skipped": 0, "errors": []},
    }

    # ──────── Puantaj sheet (3 satır per kurye) ────────
    if "Puantaj" in wb.sheetnames:
        ws = wb["Puantaj"]
        HEADER_ROW = 5  # generate ile uyumlu
        DATA_START = HEADER_ROW + 1
        # Toplam alt satırı atla — son personel satırından sonraki "Toplam"
        row_idx = DATA_START
        last_real_row = ws.max_row
        while row_idx <= last_real_row:
            tip_cell = ws.cell(row=row_idx, column=5).value
            if not tip_cell or str(tip_cell).strip() not in ("Saat", "Paket", "Durum"):
                row_idx += 1
                continue
            # Bir personel = 3 ardışık satır (Saat / Paket / Durum)
            saat_row = row_idx
            paket_row = row_idx + 1
            durum_row = row_idx + 2
            # Sanity: aşağıdaki iki satır da aynı person_code mu?
            person_code = str(ws.cell(row=saat_row, column=1).value or "").strip()
            if not person_code:
                row_idx += 3
                continue

            person = _lookup_personnel(person_code)
            if not person:
                result["puantaj"]["errors"].append(
                    f"Personel bulunamadı: {person_code} (satır {saat_row})"
                )
                row_idx += 3
                continue

            pid = int(person["id"])
            assigned_rid = person.get("assigned_restaurant_id")

            for d in range(1, n_days + 1):
                col = 5 + d
                hours_val = ws.cell(row=saat_row, column=col).value
                pkts_val = ws.cell(row=paket_row, column=col).value
                status_val = ws.cell(row=durum_row, column=col).value

                hours = float(hours_val) if hours_val not in (None, "") else 0.0
                pkts = int(pkts_val) if pkts_val not in (None, "") else 0
                status_code = (
                    str(status_val).strip().upper()[:1] if status_val else ""
                )

                # cell_type belirle
                if status_code in STATUS_CODE_MAP:
                    cell_type = STATUS_CODE_MAP[status_code]
                    coverage_type = None
                elif hours > 0 or pkts > 0:
                    cell_type = "normal"
                    coverage_type = None
                else:
                    # Tamamen boş — kayıt oluşturma
                    continue

                entry_date = f"{yyyy_i:04d}-{mm_i:02d}-{d:02d}"
                try:
                    res = upsert_cell(
                        personnel_id=pid,
                        entry_date=entry_date,
                        cell_type=cell_type,
                        worked_hours=hours,
                        package_count=pkts,
                        coverage_type=coverage_type,
                        restaurant_id=assigned_rid,
                    )
                    if res.get("action") == "created":
                        result["puantaj"]["inserted"] += 1
                    elif res.get("action") == "updated":
                        result["puantaj"]["updated"] += 1
                    else:
                        result["puantaj"]["skipped"] += 1
                except Exception as e:  # noqa: BLE001
                    result["puantaj"]["errors"].append(
                        f"{person_code} · {entry_date}: {e}"
                    )

            row_idx += 3  # sonraki personel

    # ──────── Destek sheet (Tip kolonu D/Y) ────────
    if "Destek" in wb.sheetnames:
        ws = wb["Destek"]
        # Header row 1; data row 2'den itibaren
        # Sütunlar: A=Tarih · B=Personel Kodu · C=Restoran · D=Tip ·
        #           E=Saat · F=Paket · G=Not
        for r in range(2, ws.max_row + 1):
            tarih = ws.cell(row=r, column=1).value
            kod = ws.cell(row=r, column=2).value
            rest_label = ws.cell(row=r, column=3).value
            tip = ws.cell(row=r, column=4).value
            saat = ws.cell(row=r, column=5).value
            paket = ws.cell(row=r, column=6).value
            not_text = ws.cell(row=r, column=7).value

            # Boş veya örnek satır
            kod_str = str(kod or "").strip()
            if not tarih or not kod_str:
                continue
            if kod_str.startswith("(") and kod_str.endswith(")"):
                continue  # parantezli placeholder

            try:
                tarih_str = str(tarih)[:10]  # YYYY-MM-DD
            except Exception:
                result["destek"]["errors"].append(f"Satır {r}: geçersiz tarih")
                continue

            person = _lookup_personnel(kod_str)
            if not person:
                result["destek"]["errors"].append(
                    f"Satır {r}: personel bulunamadı ({kod_str})"
                )
                continue

            rest_str = str(rest_label or "").strip()
            if rest_str.startswith("(") and rest_str.endswith(")"):
                continue  # parantezli placeholder
            rid = _lookup_restaurant(rest_str)
            if not rid:
                result["destek"]["errors"].append(
                    f"Satır {r}: restoran bulunamadı ({rest_label})"
                )
                continue

            # Tip belirle (D=Destek ücretli, Y=Yönetim ücretsiz)
            tip_code = str(tip or "").strip().upper()[:1]
            if tip_code == "Y":
                coverage = "Yönetim"
            else:
                # 'D' veya boş → varsayılan Destek
                coverage = "Destek"

            saat_v = float(saat) if saat not in (None, "") else 0.0
            pkt_v = int(paket) if paket not in (None, "") else 0
            try:
                res = upsert_cell(
                    personnel_id=int(person["id"]),
                    entry_date=tarih_str,
                    cell_type="normal",
                    worked_hours=saat_v,
                    package_count=pkt_v,
                    coverage_type=coverage,
                    restaurant_id=rid,
                    notes=str(not_text) if not_text else None,
                )
                if res.get("action") == "created":
                    result["destek"]["inserted"] += 1
                elif res.get("action") == "updated":
                    result["destek"]["updated"] += 1
                else:
                    result["destek"]["skipped"] += 1
            except Exception as e:  # noqa: BLE001
                result["destek"]["errors"].append(
                    f"Satır {r}: {e}"
                )

    return result
