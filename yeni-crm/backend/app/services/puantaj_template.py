"""Toplu puantaj Excel şablonu üretimi.

Operasyon ekibinin bir restoran (veya tüm aktif personel) için aylık puantajı
toplu olarak doldurabileceği Excel şablonu üretir.

Şablon yapısı:
  - 1 sheet: "Puantaj"
  - Üstte info satırı: dönem (Mart 2026) + restoran adı (varsa)
  - Header: Kod | Ad Soyad | Rol | Restoran | Tip | 1 | 2 | ... | N | Toplam
  - Her kurye iki satır:
      * Saat — her gün için çalışılan saat
      * Paket — her gün için teslim edilen paket
  - Hafta sonu hücreleri arka plan farklı (gri tonlu)
  - Toplam kolonunda SUM formülü
  - Açıklama satırı: "Boş hücre = gelmedi; sayı = aktif gün"

Geri dönüş: bytes (XLSX)
"""
from __future__ import annotations

import io
from calendar import monthrange
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from psycopg.rows import dict_row

from app.core.database import get_connection


TR_MONTHS = [
    "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
    "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık",
]
TR_DAYS_SHORT = ["Pzt", "Sal", "Çar", "Per", "Cum", "Cmt", "Paz"]  # Monday=0


# Çat Kapında brand palet
BRAND = "0F52BA"
BRAND_DARK = "0A3F8F"
BRAND_SOFT = "E8EFFB"
CREAM_50 = "FDFAF3"
CREAM_100 = "F5EDD8"
ROW_SAAT = "FFF7E6"
ROW_PAKET = "EFF5FB"
ROW_DURUM = "F3E8FF"  # eflatun tonu
WEEKEND_BG = "F1F4F9"
HEADER_TEXT = "FFFFFF"
TEXT = "0B0D17"
BORDER_GRAY = "ECEEF3"


def _format_period(period: str) -> str:
    y, m = period.split("-")
    return f"{TR_MONTHS[int(m) - 1]} {y}"


def _thin_side() -> Side:
    return Side(style="thin", color=BORDER_GRAY)


def generate_puantaj_template(
    period: str, restaurant_id: int | None = None,
) -> bytes:
    """Excel şablonu üret. restaurant_id verilmezse tüm aktif Kurye/Joker dahil.

    Args:
        period: 'YYYY-MM'
        restaurant_id: opsiyonel filtre — sadece bu restorana atanmış personel

    Returns:
        bytes (xlsx)
    """
    yyyy, mm = period.split("-")
    yyyy_i, mm_i = int(yyyy), int(mm)
    n_days = monthrange(yyyy_i, mm_i)[1]

    # 1) Restoran info
    rest_name = ""
    if restaurant_id:
        with get_connection() as conn:
            with conn.cursor(row_factory=dict_row) as cur:
                cur.execute(
                    "SELECT brand, branch FROM restaurants WHERE id = %s",
                    (restaurant_id,),
                )
                r = cur.fetchone()
                if r:
                    rest_name = r["brand"] or ""
                    if r.get("branch"):
                        rest_name += f" / {r['branch']}"

    # 2) Personel listesi — aktif + (restaurant_id varsa) o restorana atanmış
    with get_connection() as conn:
        with conn.cursor(row_factory=dict_row) as cur:
            sql = """
                SELECT
                    p.id, p.person_code, p.full_name, p.role,
                    p.assigned_restaurant_id,
                    r.brand AS rest_brand, r.branch AS rest_branch
                FROM personnel p
                LEFT JOIN restaurants r ON r.id = p.assigned_restaurant_id
                WHERE COALESCE(p.status, 'Aktif') = 'Aktif'
                  AND p.role IN ('Kurye', 'Joker', 'Restoran Takım Şefi',
                                 'Kaptan', 'Bölge Müdürü')
            """
            params: list = []
            if restaurant_id:
                sql += " AND p.assigned_restaurant_id = %s"
                params.append(restaurant_id)
            sql += """
                ORDER BY
                    CASE p.role
                        WHEN 'Bölge Müdürü' THEN 1
                        WHEN 'Kaptan' THEN 2
                        WHEN 'Restoran Takım Şefi' THEN 3
                        WHEN 'Joker' THEN 4
                        ELSE 5
                    END,
                    p.full_name
            """
            cur.execute(sql, tuple(params))
            personnel = cur.fetchall()

    # 3) Workbook oluştur
    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = "Puantaj"

    # ──────── INFO BARI (Row 1-3) ────────
    ws.cell(row=1, column=1, value="ÇAT KAPINDA — TOPLU PUANTAJ ŞABLONU")
    ws.cell(row=1, column=1).font = Font(
        name="Arial", size=14, bold=True, color=BRAND_DARK,
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    info_text = f"Dönem: {_format_period(period)}  ·  {n_days} gün"
    if rest_name:
        info_text += f"  ·  Restoran: {rest_name}"
    info_text += f"  ·  {len(personnel)} personel"
    ws.cell(row=2, column=1, value=info_text)
    ws.cell(row=2, column=1).font = Font(name="Arial", size=10, color="4D5468")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    # Açıklama (Row 3)
    note = (
        "Her kurye için iki satır: 'Saat' (çalışılan saat) ve 'Paket' "
        "(teslim edilen paket). Boş hücre = o gün gelmedi. Hafta sonu sütunları "
        "gri tonludur. Toplam kolonları otomatik hesaplanır."
    )
    ws.cell(row=3, column=1, value=note)
    ws.cell(row=3, column=1).font = Font(
        name="Arial", size=9, italic=True, color="8B92A7",
    )
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=8 + n_days)
    ws.cell(row=3, column=1).alignment = Alignment(
        wrap_text=True, vertical="center",
    )
    ws.row_dimensions[3].height = 30

    # ──────── HEADER (Row 5) ────────
    HEADER_ROW = 5
    headers = ["Personel Kodu", "Ad Soyad", "Rol", "Restoran", "Tip"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=HEADER_ROW, column=col_idx, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color=HEADER_TEXT)
        c.fill = PatternFill("solid", start_color=BRAND_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(
            top=_thin_side(), bottom=_thin_side(),
            left=_thin_side(), right=_thin_side(),
        )

    # Gün başlıkları (1, 2, 3, ...)
    for d in range(1, n_days + 1):
        col = 5 + d
        day_date = date(yyyy_i, mm_i, d)
        weekday = day_date.weekday()  # 0=Mon
        is_weekend = weekday >= 5  # Sat, Sun

        # Hücre: "1\nPzt"
        c = ws.cell(
            row=HEADER_ROW, column=col,
            value=f"{d}\n{TR_DAYS_SHORT[weekday]}",
        )
        c.font = Font(
            name="Arial", size=9, bold=True,
            color=HEADER_TEXT,
        )
        c.fill = PatternFill(
            "solid",
            start_color="6B7280" if is_weekend else BRAND,
        )
        c.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True,
        )
        c.border = Border(
            top=_thin_side(), bottom=_thin_side(),
            left=_thin_side(), right=_thin_side(),
        )

    # Toplam sütunu (sağ uçta)
    total_col = 5 + n_days + 1
    c = ws.cell(row=HEADER_ROW, column=total_col, value="Toplam")
    c.font = Font(name="Arial", size=10, bold=True, color=HEADER_TEXT)
    c.fill = PatternFill("solid", start_color=BRAND_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(
        top=_thin_side(), bottom=_thin_side(),
        left=_thin_side(), right=_thin_side(),
    )

    # Header yüksekliği
    ws.row_dimensions[HEADER_ROW].height = 32

    # Sütun genişlikleri
    ws.column_dimensions["A"].width = 13  # kod
    ws.column_dimensions["B"].width = 24  # ad
    ws.column_dimensions["C"].width = 14  # rol
    ws.column_dimensions["D"].width = 22  # restoran
    ws.column_dimensions["E"].width = 8   # tip (Saat/Paket)
    for d in range(1, n_days + 1):
        ws.column_dimensions[get_column_letter(5 + d)].width = 6
    ws.column_dimensions[get_column_letter(total_col)].width = 10

    # Freeze: header altı + sol kolonlar (Tip sütununa kadar)
    ws.freeze_panes = "F6"

    # ──────── PERSONEL SATIRLARI ────────
    row = HEADER_ROW + 1
    for p in personnel:
        rest_label = ""
        if p.get("rest_brand"):
            rest_label = p["rest_brand"]
            if p.get("rest_branch"):
                rest_label += f" / {p['rest_branch']}"

        # 3 satır: Saat / Paket / Durum
        for offset, label in ((0, "Saat"), (1, "Paket"), (2, "Durum")):
            rr = row + offset
            ws.cell(row=rr, column=1, value=p.get("person_code") or "")
            ws.cell(row=rr, column=2, value=p.get("full_name") or "")
            ws.cell(row=rr, column=3, value=p.get("role") or "")
            ws.cell(row=rr, column=4, value=rest_label)
            ws.cell(row=rr, column=5, value=label)

        # Stil — üç satır
        for r_off, fill_color in [(0, ROW_SAAT), (1, ROW_PAKET), (2, ROW_DURUM)]:
            rr = row + r_off
            # Sol bilgi sütunları (A-E)
            for col in range(1, 6):
                c = ws.cell(row=rr, column=col)
                c.font = Font(
                    name="Arial", size=10,
                    bold=(col in (1, 5)),
                )
                c.fill = PatternFill("solid", start_color=fill_color)
                c.alignment = Alignment(
                    horizontal="left" if col != 5 else "center",
                    vertical="center",
                )
                c.border = Border(
                    top=_thin_side(), bottom=_thin_side(),
                    left=_thin_side(), right=_thin_side(),
                )

            # Gün hücreleri (boş — kullanıcı dolduracak)
            for d in range(1, n_days + 1):
                col = 5 + d
                day_date = date(yyyy_i, mm_i, d)
                is_weekend = day_date.weekday() >= 5
                c = ws.cell(row=rr, column=col, value=None)
                c.font = Font(name="Arial", size=10)
                c.fill = PatternFill(
                    "solid",
                    start_color=WEEKEND_BG if is_weekend else "FFFFFF",
                )
                c.alignment = Alignment(
                    horizontal="center", vertical="center",
                )
                c.border = Border(
                    top=_thin_side(), bottom=_thin_side(),
                    left=_thin_side(), right=_thin_side(),
                )
                # Sayı formatı: saat ondalık, paket tam, durum text
                if r_off == 0:
                    c.number_format = "0.0"
                elif r_off == 1:
                    c.number_format = "0"
                # r_off==2 → text, format dokunma

            # Toplam (sadece Saat ve Paket satırlarında)
            first_col_letter = get_column_letter(6)
            last_col_letter = get_column_letter(5 + n_days)
            if r_off < 2:
                total_c = ws.cell(
                    row=rr, column=total_col,
                    value=f"=SUM({first_col_letter}{rr}:{last_col_letter}{rr})",
                )
                total_c.number_format = "0.0" if r_off == 0 else "0"
            else:
                # Durum satırı toplam yok
                total_c = ws.cell(row=rr, column=total_col, value="")
            total_c.font = Font(name="Arial", size=10, bold=True, color=BRAND_DARK)
            total_c.fill = PatternFill("solid", start_color=CREAM_100)
            total_c.alignment = Alignment(horizontal="center", vertical="center")
            total_c.border = Border(
                top=_thin_side(), bottom=_thin_side(),
                left=_thin_side(), right=_thin_side(),
            )

        row += 3  # 3 satır kullandık

    # ──────── ALT TOPLAM SATIRI ────────
    summary_row = row + 1
    ws.cell(row=summary_row, column=1, value="Toplam")
    ws.cell(row=summary_row, column=1).font = Font(
        name="Arial", size=11, bold=True, color=HEADER_TEXT,
    )
    ws.merge_cells(
        start_row=summary_row, start_column=1,
        end_row=summary_row, end_column=4,
    )
    ws.cell(row=summary_row, column=1).fill = PatternFill(
        "solid", start_color=BRAND_DARK,
    )
    ws.cell(row=summary_row, column=1).alignment = Alignment(
        horizontal="right", vertical="center",
    )

    ws.cell(row=summary_row, column=5, value="Saat/Paket")
    ws.cell(row=summary_row, column=5).font = Font(
        name="Arial", size=10, bold=True, color=HEADER_TEXT,
    )
    ws.cell(row=summary_row, column=5).fill = PatternFill(
        "solid", start_color=BRAND_DARK,
    )
    ws.cell(row=summary_row, column=5).alignment = Alignment(
        horizontal="center", vertical="center",
    )

    # Gün toplamları (saat + paket karışık olduğu için sadece total kolonunda)
    data_start = HEADER_ROW + 1
    data_end = row - 1  # son personel satırı
    for d in range(1, n_days + 1):
        col_letter = get_column_letter(5 + d)
        c = ws.cell(
            row=summary_row, column=5 + d,
            value=f"=SUM({col_letter}{data_start}:{col_letter}{data_end})",
        )
        c.font = Font(name="Arial", size=10, bold=True, color=HEADER_TEXT)
        c.fill = PatternFill("solid", start_color=BRAND)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = "0.0"

    # Genel toplam
    total_letter = get_column_letter(total_col)
    c = ws.cell(
        row=summary_row, column=total_col,
        value=f"=SUM({total_letter}{data_start}:{total_letter}{data_end})",
    )
    c.font = Font(name="Arial", size=11, bold=True, color=HEADER_TEXT)
    c.fill = PatternFill("solid", start_color=BRAND_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.number_format = "0.0"

    ws.row_dimensions[summary_row].height = 28

    # ──────── DESTEK SHEET ────────
    # Kuryelerin kendi restoranı dışındaki vardiyaları için ayrı kayıt.
    # Format: Tarih · Kurye kodu · Restoran (brand / branch) · Saat · Paket · Not
    support_ws = wb.create_sheet("Destek")
    support_headers = ["Tarih (YYYY-MM-DD)", "Kurye Kodu", "Restoran", "Saat", "Paket", "Not"]
    for i, h in enumerate(support_headers, start=1):
        c = support_ws.cell(row=1, column=i, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color=HEADER_TEXT)
        c.fill = PatternFill("solid", start_color=BRAND_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(
            top=_thin_side(), bottom=_thin_side(),
            left=_thin_side(), right=_thin_side(),
        )
    # Örnek satır
    support_ws.cell(row=2, column=1, value=f"{period}-15")
    support_ws.cell(row=2, column=2, value="(kurye kodu)")
    support_ws.cell(row=2, column=3, value="(restoran adı veya kodu)")
    support_ws.cell(row=2, column=4, value=11)
    support_ws.cell(row=2, column=5, value=30)
    support_ws.cell(row=2, column=6, value="örnek — bu satırı silebilirsiniz")
    for i in range(1, 7):
        support_ws.cell(row=2, column=i).font = Font(
            name="Arial", size=9, italic=True, color="8B92A7",
        )
    support_ws.column_dimensions["A"].width = 18
    support_ws.column_dimensions["B"].width = 14
    support_ws.column_dimensions["C"].width = 28
    support_ws.column_dimensions["D"].width = 10
    support_ws.column_dimensions["E"].width = 10
    support_ws.column_dimensions["F"].width = 30

    # ──────── RESTORANLAR REFERANS SHEET ────────
    # Aktif restoranların tam listesi — Destek sheet'inde restoran adı
    # yazarken kolaylık olsun, ID ve tam etiket görsünler.
    with get_connection() as conn:
        with conn.cursor(row_factory=dict_row) as cur:
            cur.execute(
                """
                SELECT id, brand, branch, pricing_model
                FROM restaurants
                WHERE COALESCE(active, 1) = 1
                ORDER BY brand, branch NULLS FIRST
                """
            )
            rest_rows = cur.fetchall()

    rest_sheet = wb.create_sheet("Restoranlar")
    rest_headers = ["ID", "Marka (Brand)", "Şube (Branch)", "Tam Etiket", "Tarife Modeli"]
    for i, h in enumerate(rest_headers, start=1):
        c = rest_sheet.cell(row=1, column=i, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color=HEADER_TEXT)
        c.fill = PatternFill("solid", start_color=BRAND_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(
            top=_thin_side(), bottom=_thin_side(),
            left=_thin_side(), right=_thin_side(),
        )

    # Açıklama satırı
    rest_sheet.cell(
        row=2, column=1,
        value=(
            "Destek sheet'inde 'Restoran' sütununa aşağıdaki tablodan kopyala — "
            "ID, Marka veya Tam Etiket olabilir."
        ),
    )
    rest_sheet.cell(row=2, column=1).font = Font(
        name="Arial", size=9, italic=True, color="8B92A7",
    )
    rest_sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    rest_sheet.cell(row=2, column=1).alignment = Alignment(
        wrap_text=True, vertical="center",
    )
    rest_sheet.row_dimensions[2].height = 24

    # Pricing model TR etiketleri
    pm_labels = {
        "hourly_only": "Saatlik",
        "hourly_plus_package": "Saat + Prim",
        "threshold_package": "Eşikli (390)",
        "fixed_monthly": "Aylık Sabit",
    }

    for i, r in enumerate(rest_rows, start=3):  # data row 3'ten itibaren
        brand = r.get("brand") or ""
        branch = r.get("branch") or ""
        tam_etiket = f"{brand} / {branch}" if branch else brand
        pm = pm_labels.get(r.get("pricing_model") or "", r.get("pricing_model") or "")

        cells = [int(r["id"]), brand, branch, tam_etiket, pm]
        is_even = (i - 3) % 2 == 0
        bg = "FFFFFF" if is_even else "F8FAFD"
        for col, val in enumerate(cells, start=1):
            c = rest_sheet.cell(row=i, column=col, value=val)
            c.font = Font(name="Arial", size=10)
            c.fill = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(
                horizontal="center" if col == 1 else "left",
                vertical="center",
            )
            c.border = Border(
                top=_thin_side(), bottom=_thin_side(),
                left=_thin_side(), right=_thin_side(),
            )

    rest_sheet.column_dimensions["A"].width = 8
    rest_sheet.column_dimensions["B"].width = 24
    rest_sheet.column_dimensions["C"].width = 24
    rest_sheet.column_dimensions["D"].width = 36
    rest_sheet.column_dimensions["E"].width = 18
    rest_sheet.row_dimensions[1].height = 22
    rest_sheet.freeze_panes = "A3"

    # ──────── KILAVUZ SHEET ────────
    guide = wb.create_sheet("Kullanım")
    guide.cell(row=1, column=1, value="ÇAT KAPINDA · TOPLU PUANTAJ ŞABLONU")
    guide.cell(row=1, column=1).font = Font(
        name="Arial", size=14, bold=True, color=BRAND_DARK,
    )
    guide_lines = [
        "",
        "1. 'Puantaj' sekmesinde her personel için ÜÇ satır vardır:",
        "   • Saat satırı: o gün çalışılan saat (örn: 11 veya 8.5)",
        "   • Paket satırı: o gün teslim edilen paket sayısı (örn: 30)",
        "   • Durum satırı: aşağıdaki TEK HARF kodları kullanın (normal çalışma için boş bırakın)",
        "",
        "2. DURUM KODLARI (Durum satırına yazılır):",
        "      G  =  Gelmedi",
        "      R  =  Raporlu",
        "      Z  =  İzin",
        "      X  =  İhbarsız çıkış",
        "      D  =  DESTEK (başka restoran kuryesi geldi — KURYE PAKET × TARİFE +",
        "             SAAT × TARİFE ücret alır, restorana fatura yansır)",
        "             → 'Destek' sekmesinde hangi restoranda çalıştığını belirtin.",
        "      Y  =  YÖNETİM kapaması (Joker / Bölge Müdürü / Kaptan / RTŞ geldi —",
        "             EKSTRA ücret YOK, sabit aylık maaşlı oldukları için sadece",
        "             operasyon kapaması olarak kayıt edilir, restorana yansımaz)",
        "",
        "3. Boş Saat/Paket + boş Durum = o gün hiç giriş yok (kayıtsız gün).",
        "   Sadece Saat ve Paket dolu, Durum boş = normal çalışma.",
        "",
        "4. Hafta sonu sütunları gri tonludur — fark etmek kolaydır.",
        "",
        "5. 'Toplam' sütunu Saat ve Paket için otomatik hesaplanır.",
        "",
        "6. 'Destek' sekmesinde restoran adı yazarken 'Restoranlar' sekmesinden "
        "kopyalayabilirsiniz (ID veya tam etiket çalışır).",
        "",
        "7. Dolu şablonu CRM'e geri yüklemek için /puantaj sayfasındaki "
        "'Excel Yükle' butonunu kullanın.",
        "",
        "Sorular için: admin@catkapinda.com",
    ]
    for i, line in enumerate(guide_lines, start=2):
        c = guide.cell(row=i, column=1, value=line)
        c.font = Font(name="Arial", size=10, color=TEXT)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    guide.column_dimensions["A"].width = 100
    for i in range(1, len(guide_lines) + 2):
        guide.row_dimensions[i].height = 18

    # ──────── EXPORT ────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
